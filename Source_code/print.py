import win32api
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from openpyxl.styles import Font, Alignment

def print_excel(data, column_names, header_text, header_font_size, column_names_font_size, width_size):
    # Create a new Workbook
    wb = Workbook()
    ws = wb.active

    # Add header with specified font size
    header_cell = ws.cell(row=1, column=1, value=header_text)
    header_cell.font = Font(size=header_font_size)
    header_cell.alignment = Alignment(horizontal='center')  # Center align the header text

    # Merge cells for header
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(column_names))
    
    # Write column names to the worksheet
    for col_name in column_names:
        cell = ws.cell(row=2, column=column_names.index(col_name) + 1, value=col_name)
        cell.font = Font(size=column_names_font_size)
        cell.alignment = Alignment(horizontal='center')  # Center align the column names

    # Write the data to the worksheet
    for row_data in data:
        # Convert phone numbers to strings
        row_data = [str(cell) for cell in row_data]
        # Add 5 empty strings between cell values to create space
        row_with_space = [val if val else '  ' for val in row_data]
        ws.append(row_with_space)

    last_column_index = len(column_names)
    for row in range(3, ws.max_row + 1):  # Start from the third row
        cell = ws.cell(row=row, column=last_column_index)
        cell.value = row - 2  # Set the value to the row number minus 2 (to start from 1)

    # Set borders for each cell to create vertical and horizontal lines
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # Skip header row
                cell.alignment = Alignment(horizontal='right')  # Right align cell content

    # Set column widths
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[chr(64 + i)].width = max_length + width_size  # Adjust width as needed

    # Save the workbook to a temporary file
    temp_file = "temp.xlsx"
    wb.save(temp_file)

    # Print the Excel file
    win32api.ShellExecute(0, "print", temp_file, None, ".", 0)